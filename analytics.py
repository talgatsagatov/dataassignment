import os
import argparse
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import text as sql_text
from config import get_engine

plt.rcParams.update({
    "figure.figsize": (10, 6),
    "axes.titlesize": 14,
    "axes.labelsize": 12,
    "legend.fontsize": 10,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "figure.dpi": 120,
})

CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"

def ensure_dirs():
    os.makedirs(CHARTS_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

SQL = {
    # 1) Pie — доля GMV по категориям
    "pie_gmv_by_category": """
        WITH cat AS (
            SELECT COALESCE(pct.product_category_name_english, p.product_category_name) AS category,
                   SUM(oi.price)::numeric(18,2) AS gmv
            FROM olist.orders o
            JOIN olist.order_items oi ON oi.order_id = o.order_id
            JOIN olist.products p     ON p.product_id = oi.product_id
            LEFT JOIN olist.product_category_name_translation pct
                   ON pct.product_category_name = p.product_category_name
            WHERE o.order_status = 'delivered'
            GROUP BY COALESCE(pct.product_category_name_english, p.product_category_name)
        ),
        ranked AS (
            SELECT category, gmv,
                   DENSE_RANK() OVER (ORDER BY gmv DESC) AS rnk
            FROM cat
        ),
        top10 AS (
            SELECT category, gmv FROM ranked WHERE rnk <= 10
        ),
        others AS (
            SELECT 'Other'::text AS category, COALESCE(SUM(gmv),0)::numeric(18,2) AS gmv
            FROM ranked WHERE rnk > 10
        )
        SELECT category, gmv FROM top10
        UNION ALL
        SELECT category, gmv FROM others
        ORDER BY gmv DESC;
    """,

    # 2) Bar — ТОП-10 штатов клиентов по GMV
    "bar_top_states_gmv": """
        SELECT c.customer_state AS state,
               SUM(oi.price)::numeric(18,2) AS gmv
        FROM olist.orders o
        JOIN olist.customers c    ON c.customer_id = o.customer_id
        JOIN olist.order_items oi ON oi.order_id  = o.order_id
        WHERE o.order_status = 'delivered'
        GROUP BY c.customer_state
        ORDER BY gmv DESC
        LIMIT 10;
    """,

    # 3) Horizontal bar — ТОП-15 продавцов по GMV
    "barh_top_sellers_gmv": """
        SELECT s.seller_id,
               SUM(oi.price)::numeric(18,2) AS gmv
        FROM olist.order_items oi
        JOIN olist.sellers s ON s.seller_id = oi.seller_id
        JOIN olist.orders  o ON o.order_id  = oi.order_id
        WHERE o.order_status = 'delivered'
        GROUP BY s.seller_id
        ORDER BY gmv DESC
        LIMIT 15;
    """,

    # 4) Line — помесячный тренд GMV
    "line_monthly_gmv": """
        SELECT DATE_TRUNC('month', o.order_purchase_timestamp)::date AS month,
               SUM(oi.price)::numeric(18,2) AS gmv
        FROM olist.orders o
        JOIN olist.order_items oi ON oi.order_id = o.order_id
        WHERE o.order_status = 'delivered'
        GROUP BY DATE_TRUNC('month', o.order_purchase_timestamp)
        ORDER BY month;
    """,

    # 5) Histogram — распределение задержки доставки (дни)
    "hist_delivery_delay_days": """
        SELECT EXTRACT(EPOCH FROM (o.order_delivered_customer_date - o.order_purchase_timestamp)) / 86400.0
                 AS delay_days
        FROM olist.orders o
        JOIN olist.customers c    ON c.customer_id = o.customer_id
        JOIN olist.order_items oi ON oi.order_id  = o.order_id
        WHERE o.order_status = 'delivered'
          AND o.order_delivered_customer_date IS NOT NULL
          AND o.order_purchase_timestamp IS NOT NULL;
    """,

    # 6) Scatter — число позиций в заказе vs сумма платежа
    "scatter_items_vs_payment": """
        WITH items AS (
          SELECT oi.order_id,
                 COUNT(*) AS n_items,
                 SUM(oi.price)::numeric(18,2) AS items_value
          FROM olist.order_items oi
          GROUP BY oi.order_id
        ),
        pays AS (
          SELECT op.order_id,
                 SUM(op.payment_value)::numeric(18,2) AS pay_total
          FROM olist.order_payments op
          GROUP BY op.order_id
        )
        SELECT o.order_id, i.n_items, i.items_value, p.pay_total
        FROM olist.orders o
        JOIN items i ON i.order_id = o.order_id
        JOIN pays  p ON p.order_id = o.order_id
        WHERE o.order_status IN ('invoiced','shipped','delivered','processing','approved');
    """,
}

def df_sql(engine, sql):
    with engine.connect() as conn:
        return pd.read_sql(sql_text(sql), conn)

def console_report(df, kind, title, path, what):
    print(f"Generated {len(df):,} rows → {kind} '{title}' → saved to {path} → {what}")

def make_pie(df, title, filename):
    label, value = df.columns[0], df.columns[1]
    series = df.set_index(label)[value]

    fig, ax = plt.subplots(figsize=(10, 7))
    wedges, texts, autotexts = ax.pie(
        series,
        startangle=90,
        counterclock=False,
        autopct=lambda pct: f"{pct:.1f}%" if pct >= 3 else "",
        pctdistance=0.8,
        labeldistance=1.1
    )
    ax.set_ylabel("")
    ax.set_title(title)

    ax.legend(
        wedges,
        [f"{idx} — {val:,.0f}" for idx, val in series.items()],
        title="Category — GMV",
        loc="center left",
        bbox_to_anchor=(1, 0.5)
    )

    out = os.path.join(CHARTS_DIR, filename)
    plt.tight_layout()
    plt.savefig(out, bbox_inches="tight")
    plt.close(fig)
    return out

def make_bar(df, title, filename, horizontal=False, x=None, y=None):
    x = x or df.columns[0]
    y = y or df.columns[1]
    ax = df.set_index(x)[y].plot.barh() if horizontal else df.set_index(x)[y].plot.bar()
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    out = os.path.join(CHARTS_DIR, filename)
    plt.tight_layout(); plt.savefig(out, bbox_inches='tight'); plt.close()
    return out

def make_line(df, title, filename, x=None, y=None):
    x = x or df.columns[0]; y = y or df.columns[1]
    if not pd.api.types.is_datetime64_any_dtype(df[x]):
        df[x] = pd.to_datetime(df[x], errors='coerce')
    ax = df.plot(x=x, y=y, kind='line')
    ax.set_title(title); ax.set_xlabel(x); ax.set_ylabel(y)
    out = os.path.join(CHARTS_DIR, filename)
    plt.tight_layout(); plt.savefig(out, bbox_inches='tight'); plt.close()
    return out

def make_hist(df, title, filename, col=None, bins=20):
    col = col or df.columns[-1]
    ax = df[col].plot(kind='hist', bins=bins)
    ax.set_title(title); ax.set_xlabel(col); ax.set_ylabel('count')
    out = os.path.join(CHARTS_DIR, filename)
    plt.tight_layout(); plt.savefig(out, bbox_inches='tight'); plt.close()
    return out

def make_scatter(df, title, filename, x=None, y=None):
    x = x or df.columns[0]; y = y or df.columns[1]
    plt.scatter(df[x], df[y])
    plt.title(title); plt.xlabel(x); plt.ylabel(y)
    out = os.path.join(CHARTS_DIR, filename)
    plt.tight_layout(); plt.savefig(out, bbox_inches='tight'); plt.close()
    return out

# plotly
def interactive_time_slider(engine):
    sql = """
        SELECT DATE_TRUNC('month', o.order_purchase_timestamp)::date AS month,
               COALESCE(pct.product_category_name_english, p.product_category_name) AS category,
               SUM(oi.price)::numeric(18,2) AS gmv
        FROM olist.orders o
        JOIN olist.order_items oi ON oi.order_id = o.order_id
        JOIN olist.products p     ON p.product_id = oi.product_id
        LEFT JOIN olist.product_category_name_translation pct
               ON pct.product_category_name = p.product_category_name
        WHERE o.order_status = 'delivered'
        GROUP BY DATE_TRUNC('month', o.order_purchase_timestamp),
                 COALESCE(pct.product_category_name_english, p.product_category_name)
        ORDER BY month, gmv DESC;
    """
    df = df_sql(engine, sql)
    if df.empty:
        print("No data for interactive slider."); 
        return

    top8 = (
        df.groupby("category", as_index=False)["gmv"].sum()
          .sort_values("gmv", ascending=False)
          .head(8)["category"].tolist()
    )
    cats = top8 + ["Other"]

    df["category"] = df["category"].where(df["category"].isin(top8), "Other")
    df = df.groupby(["month", "category"], as_index=False)["gmv"].sum()

    frames = []
    for m, g in df.groupby("month"):
        g = g.set_index("category").reindex(cats, fill_value=0).reset_index()
        g["month"] = m 
        frames.append(g)
    ready = pd.concat(frames, ignore_index=True)

    ready["month"] = pd.to_datetime(ready["month"])
    ready["month_str"] = ready["month"].dt.strftime("%Y-%m")

    fig = px.bar(
        ready,
        x="category",
        y="gmv",
        color="category",
        category_orders={"category": cats},
        animation_frame="month_str",
        title="Monthly GMV by Category (interactive)"
    )
    fig.update_layout(
        autosize=False, width=1100, height=600,
        bargap=0.25,
        margin=dict(l=40, r=220, t=60, b=80),
        legend_title_text="Category",
        xaxis=dict(title="category", tickangle=-35),
        yaxis=dict(title="gmv")
    )
    fig.show()


# export
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def export_to_excel(dfs: dict, filename: str):
    ensure_dirs(); path = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=str(sheet)[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx in range(1, ws.max_column + 1):
            if ws.max_row >= 3:
                v = ws.cell(row=2, column=col_idx).value
                if isinstance(v, (int, float)):
                    col_letter = get_column_letter(col_idx)
                    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    ws.conditional_formatting.add(rng, rule)
    wb.save(path)
    total_rows = sum(len(df) for df in dfs.values())
    print(f"Created file {os.path.basename(path)}, {len(dfs)} sheets, {total_rows:,} rows")

#Orchestrator
def run_all(engine):
    ensure_dirs(); outputs = {}

    d = df_sql(engine, SQL["pie_gmv_by_category"])
    p = make_pie(d, "GMV Share by Category", "pie_gmv_by_category.png")
    console_report(d, "pie", "GMV Share by Category", p, "Revenue distribution across categories")
    outputs["pie_gmv_by_category"] = d

    d = df_sql(engine, SQL["bar_top_states_gmv"])
    p = make_bar(d, "Top 10 States by GMV", "bar_top_states_gmv.png", x="state", y="gmv")
    console_report(d, "bar", "Top 10 States by GMV", p, "Which states drive revenue")
    outputs["bar_top_states_gmv"] = d

    d = df_sql(engine, SQL["barh_top_sellers_gmv"])
    p = make_bar(d, "Top 15 Sellers by GMV", "barh_top_sellers_gmv.png",
                 horizontal=True, x="seller_id", y="gmv")
    console_report(d, "horizontal bar", "Top 15 Sellers by GMV", p,
                   "Which sellers generate most revenue")
    outputs["barh_top_sellers_gmv"] = d

    d = df_sql(engine, SQL["line_monthly_gmv"])
    p = make_line(d, "Monthly GMV Trend", "line_monthly_gmv.png", x="month", y="gmv")
    console_report(d, "line", "Monthly GMV Trend", p, "GMV over time")
    outputs["line_monthly_gmv"] = d

    d = df_sql(engine, SQL["hist_delivery_delay_days"])
    p = make_hist(d, "Delivery Delay (days)", "hist_delivery_delay_days.png",
                  col="delay_days", bins=20)
    console_report(d, "hist", "Delivery Delay (days)", p, "Logistics performance dispersion")
    outputs["hist_delivery_delay_days"] = d

    d = df_sql(engine, SQL["scatter_items_vs_payment"])
    p = make_scatter(d, "Items per Order vs Payment", "scatter_items_vs_payment.png",
                     x="n_items", y="pay_total")
    console_report(d, "scatter", "Items vs Payment", p, "Basket size vs monetization")
    outputs["scatter_items_vs_payment"] = d

    export_to_excel(outputs, "assignment2_olist_summary.xlsx")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-all", action="store_true", help="Generate all charts + Excel export")
    parser.add_argument("--interactive", action="store_true", help="Show Plotly time slider")
    args = parser.parse_args()

    engine = get_engine()

    if args.run_all:
        run_all(engine)
    if args.interactive:
        interactive_time_slider(engine)
    if not (args.run_all or args.interactive):
        parser.print_help()
